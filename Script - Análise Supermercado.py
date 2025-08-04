import pandas as pd
from io import StringIO
from enum import Enum
import openpyxl
from openpyxl.styles import PatternFill
from datetime import date

# Dados de entrada dos preços dos produtos nos supermercados (formato CSV como string)

# Ler os dados CSV para um DataFrame do pandas
df = pd.read_csv('../precos_supermercados.csv')
# Definir a coluna 'Item' como índice do DataFrame
df = df.set_index('Item')
# Obter a lista de nomes dos supermercados (colunas do DataFrame)
todos_supermercados = df.columns.tolist()
todos_produtos = df.index.tolist()

# Definir valores mínimos de compra para cada supermercado usando Enum
class Valor_Minimo_Mercado(Enum):
    TAUSTE = None
    BARBOSA = 100
    CONFIANCA = None
    COOPSUPERMERCADO = None
    TENDAATACADO = None
    BOASUPERMERCADO = None

# Definir taxas de frete para cada supermercado usando Enum
class ValorFreteMercados(Enum):
    TAUSTE = 14.90
    BARBOSA = 16.90 + 4
    CONFIANCA = 18.90
    COOPSUPERMERCADO = 15
    TENDAATACADO = 14.90
    BOASUPERMERCADO = 15

# Criar dicionário para armazenar os valores mínimos de compra
valor_minimo = {mercado.name: mercado.value for mercado in Valor_Minimo_Mercado}
# Criar dicionário para armazenar as taxas de frete
valor_frete = {mercado.name: mercado.value for mercado in ValorFreteMercados}

def analisar_mercados_com_similaridade(data_frame, mercados_ativos, margem_similaridade=0.20):
    df_ativos = data_frame[mercados_ativos]
    menores_precos = df_ativos.min(axis=1)
    analise_detalhada = {}
    for mercado in mercados_ativos:
        preco_mercado = df_ativos[mercado]
        itens_menor_preco = []
        for item in df_ativos.index:
            if preco_mercado.loc[item] == menores_precos.loc[item]:
                count_min_price = (df_ativos.loc[item] == menores_precos.loc[item]).sum()
                if count_min_price == 1:
                    itens_menor_preco.append(item)
        analise_detalhada[mercado] = {
            'menor_preco_count': len(itens_menor_preco),
            'itens_menor_preco': itens_menor_preco
        }
    return analise_detalhada

# Analisar todos os mercados para encontrar os itens com menor preço único
analise_todos_mercados = analisar_mercados_com_similaridade(df, todos_supermercados)

# Ordenar os mercados pelo número de itens com menor preço único (em ordem decrescente)
mercados_ordenados = sorted(analise_todos_mercados.items(), key=lambda item: item[1]['menor_preco_count'], reverse=True)

# Selecionar os top 4 mercados
top_3_mercados = [mercado for mercado, detalhes in mercados_ordenados]
top_3_mercados_detalhes = []

for mercado in top_3_mercados:
    if mercado in analise_todos_mercados:
        detalhes_mercado = analise_todos_mercados[mercado]
        precos_mercado = df[mercado].to_dict()
        try:
            frete = ValorFreteMercados[f'{mercado.upper()}'].value
        except KeyError:
            frete = 0
        try:
            minimo = Valor_Minimo_Mercado[f'{mercado.upper()}'].value if Valor_Minimo_Mercado[f'{mercado.upper()}'].value is not None else 0
        except KeyError:
            minimo = 0
        top_3_mercados_detalhes.append({
            'nome': mercado,
            'precos': precos_mercado,
            'frete': frete,
            'minimo': minimo,
            'itens_menor_preco': detalhes_mercado['itens_menor_preco']
        })
        print(f"Melhor mercado encontrado: {mercado}")
        print(f"Itens com menor preço: {detalhes_mercado['itens_menor_preco']}")
        print(f"Valor do frete para {mercado}: {frete}")

if top_3_mercados_detalhes:
    nome_arquivo_excel = 'analise_supermercados.xlsx'
    # Escrever os dados para o Excel
    with pd.ExcelWriter('./Resultados/' + nome_arquivo_excel.replace('.xlsx', f'_{date.today().strftime("%Y-%m-%d")}.xlsx'), engine='openpyxl') as writer:
        cor_verde = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid') # Verde
        menor_preco_por_produto = df.min(axis=1)

        for i, detalhes in enumerate(top_3_mercados_detalhes):
            nome_planilha = f"TOP {i+1} {detalhes['nome']}"
            supermercado_atual = detalhes['nome']
            precos_mercado_atual = df[supermercado_atual].to_dict()

            data = {'Produto': todos_produtos, supermercado_atual: [precos_mercado_atual.get(produto) for produto in todos_produtos]}
            df_planilha = pd.DataFrame(data)
            df_planilha.to_excel(writer, sheet_name=nome_planilha, index=False)

    # Agora, abrir o arquivo e adicionar formatação
    workbook = openpyxl.load_workbook('./Resultados/' + nome_arquivo_excel.replace('.xlsx', f'_{date.today().strftime("%Y-%m-%d")}.xlsx'))
    cor_verde = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid') # Verde
    menor_preco_por_produto = df.min(axis=1)

    for i, detalhes in enumerate(top_3_mercados_detalhes):
        nome_planilha = f"TOP {i+1} {detalhes['nome']}"
        supermercado_atual = detalhes['nome']
        if nome_planilha in workbook.sheetnames:
            worksheet = workbook[nome_planilha]
            valor_total_baratos = 0
            for row in range(2, len(todos_produtos) + 2):
                produto = worksheet['A' + str(row)].value
                preco_no_mercado = df.loc[produto, supermercado_atual]
                preco_minimo = menor_preco_por_produto.loc[produto]
                if preco_no_mercado == preco_minimo:
                    valor_total_baratos += preco_no_mercado
                    cell_produto = worksheet['A' + str(row)]
                    cell_produto.fill = cor_verde

            # Adicionar as linhas de Frete, Valor Mínimo e Valor Total ao final de cada planilha
            ultima_linha = worksheet.max_row + 1
            worksheet['A' + str(ultima_linha)] = 'Frete'
            worksheet['B' + str(ultima_linha)] = detalhes['frete']
            ultima_linha += 1
            worksheet['A' + str(ultima_linha)] = 'Valor Mínimo'
            worksheet['B' + str(ultima_linha)] = detalhes['minimo']
            ultima_linha += 1
            worksheet['A' + str(ultima_linha)] = 'Valor Total'
            valor_total_itens = df[supermercado_atual].sum()
            valor_total = valor_total_itens + detalhes['frete']
            worksheet['B' + str(ultima_linha)] = valor_total
            ultima_linha += 1
            worksheet['A' + str(ultima_linha)] = 'Total Baratos + Frete'
            cell_total_baratos_label = worksheet['A' + str(ultima_linha)]
            cell_total_baratos_label.fill = cor_verde
            worksheet['B' + str(ultima_linha)] = valor_total_baratos + detalhes['frete']
            cell_total_baratos_value = worksheet['B' + str(ultima_linha)]


    workbook.save('./Resultados/' + nome_arquivo_excel.replace('.xlsx', f'_{date.today().strftime("%Y_%m_%d")}.xlsx'))
    # print(f"\nOs dados dos top 4 supermercados foram exportados para o arquivo './Resultados/{nome_arquivo_excel.replace('.xlsx', f'_{date.today().strftime('%Y-%m-%d')}.xlsx')}'.")
    print("Os produtos com o menor preço em cada supermercado foram pintados de verde.")

else:
    print("\nNão foram encontrados mercados suficientes para exportar para o Excel.")